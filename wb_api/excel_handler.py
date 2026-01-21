"""Модуль для работы с Excel таблицами"""
import pandas as pd
import numpy as np
import logging
from datetime import datetime
from typing import Dict, List, Optional
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Класс для работы с Excel файлами"""
    
    def __init__(self, file_path: Optional[str] = None, template_path: Optional[str] = None):
        """
        Инициализация обработчика Excel
        
        Args:
            file_path: Путь к файлу Excel для работы (если None, будет создан новый с датой и временем)
            template_path: Путь к шаблону таблицы (по умолчанию wb_data.xlsx)
        """
        self.template_path = template_path or "wb_data.xlsx"
        
        # Если file_path не указан, создаем новый файл с датой и временем
        if file_path is None:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_path = f"wb_data_{timestamp}.xlsx"
        
        self.file_path = file_path
        self._last_date = None  # Сохраняем последнюю дату для использования при записи
        
        # Создаем директорию для файла, если её нет
        file_dir = os.path.dirname(self.file_path)
        if file_dir and not os.path.exists(file_dir):
            os.makedirs(file_dir, exist_ok=True)
            logger.info(f"Создана директория: {file_dir}")
        
        self._create_from_template()
    
    def _get_template_columns(self) -> List[str]:
        """
        Получает список колонок из шаблона таблицы
        
        Returns:
            Список названий колонок
        """
        if os.path.exists(self.template_path):
            try:
                # Пробуем разные варианты чтения заголовков
                for header_row in [0, 3]:
                    try:
                        df_template = pd.read_excel(self.template_path, engine='openpyxl', header=header_row)
                        # Проверяем, что это действительно заголовки (не числовые значения)
                        first_col = str(df_template.columns[0])
                        if "акк" in first_col.lower() or "дата" in first_col.lower():
                            columns = [str(col).strip() for col in df_template.columns if pd.notna(col)]
                            logger.info(f"Загружены колонки из шаблона (строка {header_row}): {len(columns)} колонок")
                            return columns
                    except:
                        continue
            except Exception as e:
                logger.warning(f"Не удалось прочитать шаблон {self.template_path}: {str(e)}")
        
        # Если шаблон не найден, используем стандартные колонки
        return [
            "акк/дата",
            "ставка/реклама",
            "комментарий",
            "Касса месяц",
            "Касса день",
            "показы/",
            "клики",
            "CTR",
            "цена клика",
            "расход АУКЦ",
            "цена клика АРК",
            "расход АРК",
            "перешли в карточку",
            "корзин",
            "заказы",
            "хран день",
            "М с 1 шт\nнаши данные",
            "итого прибыль",
            "ЦЕНА \nтовара",
            "остаток товара на складе"
        ]
    
    def _create_from_template(self):
        """Создает новый файл Excel на основе шаблона с сохранением форматирования"""
        if not os.path.exists(self.template_path):
            logger.warning(f"Шаблон {self.template_path} не найден, создаю файл со стандартными колонками")
            columns = self._get_template_columns()
            df = pd.DataFrame(columns=columns)
            df.to_excel(self.file_path, index=False, engine='openpyxl')
            logger.info(f"Создан новый Excel файл: {self.file_path} с {len(columns)} колонками")
            return
        
        try:
            # Копируем шаблон в новый файл (это сохраняет все форматирование, размеры столбцов, цвета)
            import shutil
            shutil.copy2(self.template_path, self.file_path)
            logger.info(f"Создан новый файл {self.file_path} на основе шаблона {self.template_path}")
            
            # Проверяем, что файл можно прочитать
            try:
                df = pd.read_excel(self.file_path, engine='openpyxl', header=0)
                logger.info(f"Файл успешно создан, колонок: {len(df.columns)}")
            except Exception as e:
                logger.warning(f"Не удалось прочитать созданный файл: {str(e)}, создаю пустой")
                columns = self._get_template_columns()
                df = pd.DataFrame(columns=columns)
                df.to_excel(self.file_path, index=False, engine='openpyxl')
        except Exception as e:
            logger.error(f"Ошибка при создании файла из шаблона: {str(e)}")
            # Fallback: создаем пустой файл
            columns = self._get_template_columns()
            df = pd.DataFrame(columns=columns)
            df.to_excel(self.file_path, index=False, engine='openpyxl')
            logger.info(f"Создан пустой Excel файл: {self.file_path}")
    
    def read_data(self) -> pd.DataFrame:
        """
        Читает данные из Excel файла
        
        Returns:
            DataFrame с данными
        """
        try:
            # Проверяем, существует ли файл и является ли он валидным Excel файлом
            if not os.path.exists(self.file_path):
                logger.warning(f"Файл {self.file_path} не существует, создаю новый")
                self._create_from_template()
                return pd.read_excel(self.file_path, engine='openpyxl', header=0)
            
            # Проверяем размер файла (пустой файл может быть поврежден)
            if os.path.getsize(self.file_path) == 0:
                logger.warning(f"Файл {self.file_path} пустой, пересоздаю")
                os.remove(self.file_path)
                self._create_from_template()
                return pd.read_excel(self.file_path, engine='openpyxl', header=0)
            
            # Пытаемся прочитать файл (заголовки в первой строке)
            return pd.read_excel(self.file_path, engine='openpyxl', header=0)
            
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
            # Если файл поврежден, пересоздаем его
            logger.info(f"Пересоздаю файл {self.file_path}")
            if os.path.exists(self.file_path):
                try:
                    os.remove(self.file_path)
                except:
                    pass
            self._create_from_template()
            return pd.read_excel(self.file_path, engine='openpyxl', header=0)
    
    def write_data(self, df: pd.DataFrame):
        """
        Записывает данные в Excel файл с сохранением форматирования шаблона
        
        Args:
            df: DataFrame для записи
        """
        # Используем openpyxl напрямую для сохранения форматирования
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Находим строку для вставки данных (после заголовков)
            # Обычно заголовки в первой строке, данные начинаются со второй
            start_row = 2  # Начинаем со второй строки (после заголовка)
            
            # Очищаем старые данные (кроме заголовков)
            if ws.max_row >= start_row:
                ws.delete_rows(start_row, ws.max_row - start_row + 1)
            
            # Получаем маппинг колонок DataFrame на колонки Excel
            df_columns = list(df.columns)
            excel_columns = {}
            
            # Находим соответствие колонок по заголовкам (нормализуем названия)
            for col_idx, cell in enumerate(ws[1], 1):
                col_name = str(cell.value).strip() if cell.value else ""
                # Нормализуем названия (убираем переносы строк, лишние пробелы)
                col_name_normalized = col_name.replace('\n', ' ').replace('\r', ' ').strip()
                
                # Ищем соответствие в DataFrame
                for df_col in df_columns:
                    df_col_normalized = str(df_col).replace('\n', ' ').replace('\r', ' ').strip()
                    # Точное совпадение или частичное
                    if (col_name_normalized == df_col_normalized or 
                        col_name_normalized.lower() == df_col_normalized.lower() or
                        col_name in str(df_col) or str(df_col) in col_name):
                        excel_columns[df_col] = col_idx
                        break
            
            # Записываем данные построчно
            for df_idx, (_, row_data) in enumerate(df.iterrows()):
                excel_row = start_row + df_idx
                
                # Сначала обрабатываем дату отдельно, чтобы гарантировать её запись
                for col_idx in range(1, ws.max_column + 1):
                    excel_col_name = str(ws.cell(row=1, column=col_idx).value).strip() if ws.cell(row=1, column=col_idx).value else ""
                    if "акк/дата" in excel_col_name.lower() or "акк/дата" in excel_col_name:
                        # Нашли колонку даты, записываем значение из DataFrame
                        for df_col in df.columns:
                            if "акк/дата" in str(df_col).lower() or "акк/дата" in str(df_col):
                                value = row_data.get(df_col, "")
                                cell = ws.cell(row=excel_row, column=col_idx)
                                
                                # Копируем форматирование
                                template_row = 2 if ws.max_row >= 2 else 1
                                template_cell = ws.cell(row=template_row, column=col_idx)
                                if template_cell.has_style:
                                    from copy import copy
                                    cell.font = copy(template_cell.font) if template_cell.font else None
                                    cell.fill = copy(template_cell.fill) if template_cell.fill else None
                                    cell.border = copy(template_cell.border) if template_cell.border else None
                                    cell.alignment = copy(template_cell.alignment) if template_cell.alignment else None
                                    cell.number_format = template_cell.number_format
                                
                                # Записываем дату - если в DataFrame NaN, используем сохраненную дату
                                if pd.notna(value) and value != "" and value != np.nan and str(value).strip() != "":
                                    cell.value = str(value)
                                elif self._last_date and df_idx == 0:  # Только для первой строки (новой)
                                    # Если дата потерялась в DataFrame, используем сохраненную дату
                                    logger.info(f"Восстанавливаю дату '{self._last_date}' из сохраненного значения для строки {excel_row}")
                                    cell.value = str(self._last_date)
                                else:
                                    logger.debug(f"Дата в DataFrame пустая для строки {excel_row}, оставляю пустой")
                                    cell.value = None
                                break
                        break
                
                # Затем обрабатываем остальные колонки
                for col_name, col_idx in excel_columns.items():
                    # Пропускаем дату, так как уже обработали
                    if "акк/дата" in str(col_name).lower() or "акк/дата" in str(col_name):
                        continue
                    
                    value = row_data.get(col_name, "")
                    cell = ws.cell(row=excel_row, column=col_idx)
                    
                    # Копируем форматирование из шаблона (берем стиль из второй строки, если есть, иначе из заголовка)
                    template_row = 2 if ws.max_row >= 2 else 1
                    template_cell = ws.cell(row=template_row, column=col_idx)
                    
                    if template_cell.has_style:
                        from copy import copy
                        cell.font = copy(template_cell.font) if template_cell.font else None
                        cell.fill = copy(template_cell.fill) if template_cell.fill else None
                        cell.border = copy(template_cell.border) if template_cell.border else None
                        cell.alignment = copy(template_cell.alignment) if template_cell.alignment else None
                        cell.number_format = template_cell.number_format
                    
                    # Устанавливаем значение
                    if pd.notna(value) and value != "" and value != np.nan:
                        # Преобразуем типы для числовых значений
                        try:
                            if isinstance(value, (int, float)):
                                cell.value = value
                            elif isinstance(value, str) and (value.replace('.', '').replace('-', '').replace('e', '').replace('E', '').isdigit() or value == ''):
                                if value == '':
                                    cell.value = None
                                else:
                                    cell.value = float(value) if '.' in value or 'e' in value.lower() else int(value)
                            else:
                                cell.value = str(value)
                        except:
                            cell.value = str(value)
            
            wb.save(self.file_path)
            logger.info(f"Данные записаны в {self.file_path} с сохранением форматирования")
            
        except Exception as e:
            logger.warning(f"Не удалось сохранить с форматированием: {str(e)}, использую pandas")
            # Fallback на pandas
            df.to_excel(self.file_path, index=False, engine='openpyxl')
    
    def _get_manual_fields(self) -> List[str]:
        """
        Возвращает список полей, которые вносятся вручную (зеленые поля)
        Эти поля не должны перезаписываться данными из API
        
        Returns:
            Список названий колонок для ручного ввода
        """
        return [
            "ставка/реклама",
            "комментарий",
            "Касса месяц",
            "М с 1 шт\nнаши данные",
            "М с 1 шт наши данные"
        ]
    
    def _is_manual_field(self, column_name: str) -> bool:
        """
        Проверяет, является ли поле ручным (зеленым)
        
        Args:
            column_name: Название колонки
            
        Returns:
            True если поле ручное, False если берется из API
        """
        manual_fields = self._get_manual_fields()
        column_str = str(column_name)
        
        # Проверяем точное совпадение или частичное для полей с переносами строк
        for manual_field in manual_fields:
            if manual_field in column_str or column_str in manual_field:
                return True
        
        # Специальные проверки
        if "ставка" in column_str.lower() and "реклама" in column_str.lower():
            return True
        if "комментарий" in column_str.lower():
            return True
        if "касса месяц" in column_str.lower():
            return True
        if "м с 1 шт" in column_str.lower() and "наши данные" in column_str.lower():
            return True
        
        return False
    
    def add_daily_data(self, date: str, wb_data: Dict, manual_data: Optional[Dict] = None):
        """
        Добавляет или обновляет дневные данные в таблице.
        Зеленые поля (ручной ввод) сохраняются, если уже заполнены.
        Остальные поля обновляются данными из API WB.
        
        Args:
            date: Дата в формате DD.MM.YYYY
            wb_data: Данные из API Wildberries
            manual_data: Данные для ручного ввода (зеленые поля) - опционально
        """
        # Сохраняем дату для использования при записи
        self._last_date = date
        """
        Добавляет или обновляет дневные данные в таблице.
        Зеленые поля (ручной ввод) сохраняются, если уже заполнены.
        Остальные поля обновляются данными из API WB.
        
        Args:
            date: Дата в формате DD.MM.YYYY
            wb_data: Данные из API Wildberries
            manual_data: Данные для ручного ввода (зеленые поля) - опционально
        """
        df = self.read_data()
        
        # Проверяем, существует ли уже строка с этой датой
        date_mask = df["акк/дата"].astype(str).str.contains(date.split('.')[0], na=False)
        existing_row_idx = None
        if date_mask.any():
            existing_row_idx = date_mask.idxmax()
            logger.debug(f"Найдена существующая строка для даты {date} (индекс {existing_row_idx}), обновляю данные из API")
            logger.debug(f"Текущие данные строки: {df.loc[existing_row_idx].to_dict()}")
            logger.debug(f"Новые данные из API: {wb_data}")
        
        if existing_row_idx is not None:
            # Обновляем существующую строку
            row = df.loc[existing_row_idx]
            
            # Сохраняем зеленые поля (ручной ввод), если они уже заполнены
            for col in df.columns:
                if self._is_manual_field(col):
                    # Не перезаписываем зеленые поля, если они уже заполнены
                    if pd.notna(row[col]) and str(row[col]).strip() != "":
                        continue  # Оставляем существующее значение
                    # Заполняем только если переданы новые данные вручную
                    if manual_data:
                        for manual_key in manual_data:
                            if manual_key in str(col) or str(col) in manual_key:
                                df.at[existing_row_idx, col] = manual_data[manual_key]
                                break
            
            # Обновляем только не зеленые поля данными из API
            self._update_wb_fields(df, existing_row_idx, wb_data)
        else:
            # Создаем новую строку
            new_row = {col: "" for col in df.columns}
            # Убеждаемся, что дата записывается как строка
            # Ищем правильное название колонки даты в DataFrame
            date_col_name = None
            for col in df.columns:
                if "акк/дата" in str(col).lower() or "акк/дата" in str(col):
                    date_col_name = col
                    break
            
            if date_col_name:
                new_row[date_col_name] = str(date) if date else ""
                logger.debug(f"Дата '{date}' записана в колонку '{date_col_name}'")
            else:
                logger.warning(f"Колонка 'акк/дата' не найдена в DataFrame. Колонки: {list(df.columns)[:5]}")
            
            # Заполняем зеленые поля (ручной ввод) только если переданы
            if manual_data:
                for col in df.columns:
                    if self._is_manual_field(col):
                        for manual_key, manual_value in manual_data.items():
                            if manual_key in str(col) or str(col) in manual_key:
                                new_row[col] = manual_value
                                break
            
            # Заполняем данные из API WB (не зеленые поля)
            self._fill_wb_fields(new_row, df.columns, wb_data)
            
            # Проверяем, что дата не потерялась перед добавлением в DataFrame
            if date_col_name and new_row.get(date_col_name) != str(date):
                logger.warning(f"Дата потерялась! Было: '{date}', стало: '{new_row.get(date_col_name)}'")
                new_row[date_col_name] = str(date) if date else ""
            
            # Добавляем новую строку
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        
        self.write_data(df)
    
    def _fill_wb_fields(self, row: Dict, columns: List, wb_data: Dict):
        """
        Заполняет поля данными из API WB (не зеленые поля)
        
        Args:
            row: Словарь для заполнения
            columns: Список колонок DataFrame
            wb_data: Данные из API Wildberries
        """
        for col in columns:
            if self._is_manual_field(col):
                continue  # Пропускаем зеленые поля
            
            col_str = str(col)
            value = ""
            
            # Маппинг данных из API на колонки таблицы
            if "касса день" in col_str.lower():
                value = wb_data.get("Касса день", "")
            elif "показы" in col_str.lower():
                value = wb_data.get("показы", "")
            elif "клики" in col_str.lower() and "клика" not in col_str.lower():
                value = wb_data.get("клики", "")
            elif col_str.strip() == "CTR":
                value = wb_data.get("CTR", "")
            elif "цена клика" in col_str.lower() and "аукц" in col_str.lower():
                value = wb_data.get("цена клика АУКЦ", "")
            elif "цена клика" in col_str.lower() and "арк" in col_str.lower():
                value = wb_data.get("цена клика АРК", "")
            elif "цена клика" in col_str.lower() and "аукц" not in col_str.lower() and "арк" not in col_str.lower():
                # Просто "цена клика" без АУКЦ/АРК
                value = wb_data.get("цена клика", "")
            elif "расход аукц" in col_str.lower():
                value = wb_data.get("расход АУКЦ", "")
            elif "расход арк" in col_str.lower():
                value = wb_data.get("расход АРК", "")
            elif "перешли в карточку" in col_str.lower():
                value = wb_data.get("перешли в карточку", "")
            elif "корзин" in col_str.lower():
                value = wb_data.get("корзин", "")
            elif "заказы" in col_str.lower():
                value = wb_data.get("заказы", "")
            elif "хран день" in col_str.lower():
                value = wb_data.get("хран день", "")
            elif "итого прибыль" in col_str.lower():
                value = wb_data.get("итого прибыль", "")
            elif "цена" in col_str.lower() and "товара" in col_str.lower():
                value = wb_data.get("ЦЕНА товара", "")
            elif "остаток" in col_str.lower() and "складе" in col_str.lower():
                value = wb_data.get("остаток товара на складе", wb_data.get("остаток товар", ""))
            
            # Устанавливаем значение (пустая строка для текстовых полей, NaN для числовых)
            row[col] = value if value != "" else np.nan
    
    def _update_wb_fields(self, df: pd.DataFrame, row_idx: int, wb_data: Dict):
        """
        Обновляет только не зеленые поля в существующей строке данными из API
        
        Args:
            df: DataFrame для обновления
            row_idx: Индекс строки для обновления
            wb_data: Данные из API Wildberries
        """
        updated_fields = []
        
        # Вспомогательная функция для безопасного присвоения значений
        def safe_set_value(value, column_name):
            """Безопасно устанавливает значение с учетом типа колонки"""
            if value == "" or value is None:
                return np.nan
            # Если колонка числовая, конвертируем в число
            if pd.api.types.is_numeric_dtype(df[column_name]):
                try:
                    num_value = float(value) if value != "" else np.nan
                    # Если колонка int64, но значение float - конвертируем колонку в float64
                    if pd.api.types.is_integer_dtype(df[column_name]) and not pd.isna(num_value) and num_value != int(num_value):
                        # Меняем тип колонки на float64 для поддержки десятичных значений
                        df[column_name] = df[column_name].astype('float64')
                    return num_value
                except (ValueError, TypeError):
                    return np.nan
            return value
        
        for col in df.columns:
            if self._is_manual_field(col):
                continue  # Пропускаем зеленые поля
            
            col_str = str(col)
            old_value = df.at[row_idx, col]
            new_value = None
            
            # Маппинг данных из API на колонки таблицы
            if "касса день" in col_str.lower():
                new_value = safe_set_value(wb_data.get("Касса день", ""), col)
            elif "показы" in col_str.lower():
                new_value = safe_set_value(wb_data.get("показы", ""), col)
            elif "клики" in col_str.lower() and "клика" not in col_str.lower():
                new_value = safe_set_value(wb_data.get("клики", ""), col)
            elif col_str.strip() == "CTR":
                new_value = safe_set_value(wb_data.get("CTR", ""), col)
            elif "цена клика" in col_str.lower() and "аукц" in col_str.lower():
                new_value = safe_set_value(wb_data.get("цена клика АУКЦ", ""), col)
            elif "цена клика" in col_str.lower() and "арк" in col_str.lower():
                new_value = safe_set_value(wb_data.get("цена клика АРК", ""), col)
            elif "цена клика" in col_str.lower() and "аукц" not in col_str.lower() and "арк" not in col_str.lower():
                # Просто "цена клика" без АУКЦ/АРК
                new_value = safe_set_value(wb_data.get("цена клика", ""), col)
            elif "расход аукц" in col_str.lower():
                new_value = safe_set_value(wb_data.get("расход АУКЦ", ""), col)
            elif "расход арк" in col_str.lower():
                new_value = safe_set_value(wb_data.get("расход АРК", ""), col)
            elif "перешли в карточку" in col_str.lower():
                new_value = safe_set_value(wb_data.get("перешли в карточку", ""), col)
            elif "корзин" in col_str.lower():
                new_value = safe_set_value(wb_data.get("корзин", ""), col)
            elif "заказы" in col_str.lower():
                new_value = safe_set_value(wb_data.get("заказы", ""), col)
            elif "хран день" in col_str.lower():
                new_value = safe_set_value(wb_data.get("хран день", ""), col)
            elif "итого прибыль" in col_str.lower():
                new_value = safe_set_value(wb_data.get("итого прибыль", ""), col)
            elif "цена" in col_str.lower() and "товара" in col_str.lower():
                new_value = safe_set_value(wb_data.get("ЦЕНА товара", ""), col)
            elif "остаток" in col_str.lower():
                new_value = safe_set_value(wb_data.get("остаток товар", ""), col)
            
            if new_value is not None:
                df.at[row_idx, col] = new_value
                if pd.notna(new_value) and (pd.isna(old_value) or str(old_value) != str(new_value)):
                    updated_fields.append(f"{col}: {old_value} → {new_value}")
        
        if updated_fields:
            logger.info(f"Обновлено полей: {len(updated_fields)}. Изменения: {', '.join(updated_fields[:10])}{'...' if len(updated_fields) > 10 else ''}")
        else:
            logger.debug(f"Нет изменений для обновления (данные идентичны или пустые)")
    
    def update_row(self, date: str, updates: Dict):
        """
        Обновляет строку с указанной датой
        
        Args:
            date: Дата для поиска строки
            updates: Словарь с обновлениями
        """
        df = self.read_data()
        mask = df["акк/дата"] == date
        if mask.any():
            for key, value in updates.items():
                if key in df.columns:
                    df.loc[mask, key] = value
            self.write_data(df)
        else:
            raise ValueError(f"Строка с датой {date} не найдена")

